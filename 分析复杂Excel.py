"""
分析复杂Excel文件结构
专门处理多层表头、复杂格式的电力行业Excel文件
"""

import pandas as pd
import openpyxl
from pathlib import Path
import numpy as np

def analyze_excel_structure(file_path):
    """深度分析Excel文件结构"""
    print(f"=== 分析文件: {file_path} ===\n")
    
    # 1. 使用openpyxl分析工作表信息
    print("📋 工作表信息:")
    try:
        wb = openpyxl.load_workbook(file_path)
        print(f"工作表数量: {len(wb.sheetnames)}")
        for i, sheet_name in enumerate(wb.sheetnames):
            print(f"  {i+1}. {sheet_name}")
        print()
        
        # 分析第一个工作表的详细结构
        first_sheet = wb[wb.sheetnames[0]]
        print(f"🔍 详细分析工作表: {wb.sheetnames[0]}")
        print(f"最大行数: {first_sheet.max_row}")
        print(f"最大列数: {first_sheet.max_column}")
        print()
        
        # 查看前20行的内容
        print("📊 前20行内容预览:")
        for row in range(1, min(21, first_sheet.max_row + 1)):
            row_data = []
            for col in range(1, min(11, first_sheet.max_column + 1)):  # 只看前10列
                cell = first_sheet.cell(row=row, column=col)
                value = cell.value
                if value is None:
                    value = ""
                elif isinstance(value, (int, float)):
                    value = str(value)
                else:
                    value = str(value)[:20]  # 截断长文本
                row_data.append(value)
            
            print(f"第{row:2d}行: {' | '.join(f'{v:15s}' for v in row_data[:8])}")
        
        print("\n" + "="*100)
        
    except Exception as e:
        print(f"openpyxl分析失败: {e}")
    
    # 2. 使用pandas尝试不同的读取方式
    print("\n📈 pandas读取尝试:")
    
    # 方式1：默认读取
    try:
        print("\n🔸 方式1 - 默认读取:")
        df1 = pd.read_excel(file_path)
        print(f"形状: {df1.shape}")
        print("列名:")
        for i, col in enumerate(df1.columns):
            print(f"  {i+1}. {col}")
        print("\n前5行:")
        print(df1.head().to_string())
    except Exception as e:
        print(f"默认读取失败: {e}")
    
    # 方式2：跳过前几行
    for skip_rows in [0, 1, 2, 3, 4, 5]:
        try:
            print(f"\n🔸 方式2 - 跳过前{skip_rows}行:")
            df2 = pd.read_excel(file_path, skiprows=skip_rows)
            print(f"形状: {df2.shape}")
            if not df2.empty:
                print("列名:")
                for i, col in enumerate(df2.columns):
                    print(f"  {i+1}. {col}")
                
                # 查找包含"202"或"主变"的数据
                print("\n🔍 搜索关键数据:")
                found_data = False
                for col in df2.columns:
                    if df2[col].dtype == 'object':  # 文本列
                        mask = df2[col].astype(str).str.contains('202|主变', na=False, case=False)
                        if mask.any():
                            found_rows = df2[mask]
                            print(f"在列 '{col}' 中找到包含'202'或'主变'的数据:")
                            print(found_rows.to_string())
                            found_data = True
                            break
                
                if not found_data:
                    print("未找到包含'202'或'主变'的数据")
                    
                break  # 找到有效数据就停止
        except Exception as e:
            print(f"跳过{skip_rows}行读取失败: {e}")
    
    # 3. 分析合并单元格
    print(f"\n🔗 合并单元格分析:")
    try:
        wb = openpyxl.load_workbook(file_path)
        first_sheet = wb[wb.sheetnames[0]]
        merged_ranges = list(first_sheet.merged_cells.ranges)
        print(f"合并单元格数量: {len(merged_ranges)}")
        if merged_ranges:
            print("合并单元格列表:")
            for i, merged_range in enumerate(merged_ranges[:10]):  # 只显示前10个
                print(f"  {i+1}. {merged_range}")
                
    except Exception as e:
        print(f"合并单元格分析失败: {e}")

def suggest_optimization(file_path):
    """基于分析结果提出优化建议"""
    print(f"\n🚀 优化建议:")
    print("="*60)
    
    suggestions = [
        "1. 📋 多表头处理:",
        "   - 使用 header=[0,1] 参数处理多层表头",
        "   - 或手动指定表头行位置",
        "",
        "2. �� 数据定位优化:",
        "   - 使用 skiprows 跳过标题和说明行",
        "   - 自动检测数据开始行",
        "",
        "3. �� 智能字段匹配:",
        "   - 支持模糊匹配字段名",
        "   - 处理合并单元格产生的空列名",
        "",
        "4. 📊 数据清洗增强:",
        "   - 自动清理空行和空列",
        "   - 处理合并单元格的数据展开",
        "",
        "5. 🎨 用户界面优化:",
        "   - 预览文件结构",
        "   - 可视化选择数据区域",
        "   - 智能推荐表头行"
    ]
    
    for suggestion in suggestions:
        print(suggestion)

if __name__ == "__main__":
    file_path = "4月东宇母线不平衡率 .xlsx"
    if Path(file_path).exists():
        analyze_excel_structure(file_path)
        suggest_optimization(file_path)
    else:
        print(f"文件不存在: {file_path}")
