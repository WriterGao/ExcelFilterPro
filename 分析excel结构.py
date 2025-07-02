"""
分析复杂Excel文件结构
"""

import pandas as pd
import openpyxl
from pathlib import Path

def analyze_excel_structure(file_path):
    """分析Excel文件结构"""
    print(f"=== 分析文件: {file_path} ===")
    
    # 使用openpyxl分析原始结构
    print("\n📋 工作表信息:")
    wb = openpyxl.load_workbook(file_path)
    for i, sheet_name in enumerate(wb.sheetnames):
        sheet = wb[sheet_name]
        print(f"  {i+1}. {sheet_name} - 最大行数: {sheet.max_row}, 最大列数: {sheet.max_column}")
    
    # 分析第一个工作表的详细结构
    first_sheet = wb.worksheets[0]
    print(f"\n📊 分析工作表: {first_sheet.title}")
    
    # 检查前10行的结构
    print("\n🔍 前10行数据结构:")
    for row_num in range(1, min(11, first_sheet.max_row + 1)):
        row_data = []
        for col_num in range(1, min(21, first_sheet.max_column + 1)):  # 前20列
            cell = first_sheet.cell(row=row_num, column=col_num)
            value = str(cell.value) if cell.value is not None else ""
            if len(value) > 15:
                value = value[:12] + "..."
            row_data.append(value)
        
        # 显示非空列
        non_empty = [f"列{i+1}:{val}" for i, val in enumerate(row_data) if val.strip()]
        if non_empty:
            print(f"  第{row_num}行: {', '.join(non_empty[:8])}")  # 只显示前8个非空列
    
    # 使用pandas尝试不同的读取方式
    print(f"\n📖 pandas读取测试:")
    
    # 方式1: 默认读取
    try:
        df1 = pd.read_excel(file_path)
        print(f"  方式1 (默认): 形状 {df1.shape}, 列名: {list(df1.columns)[:5]}")
        print(f"    前几行数据:")
        for i, row in df1.head(3).iterrows():
            print(f"      第{i+1}行: {dict(row)}")
    except Exception as e:
        print(f"  方式1失败: {e}")
    
    # 方式2: 跳过前几行
    for skip_rows in [1, 2, 3, 4, 5]:
        try:
            df_skip = pd.read_excel(file_path, skiprows=skip_rows)
            if not df_skip.empty and len(df_skip.columns) > 3:
                print(f"  方式2 (跳过{skip_rows}行): 形状 {df_skip.shape}, 列名: {list(df_skip.columns)[:5]}")
                if not df_skip.empty:
                    # 找到包含"202"或"主变"的行
                    sample_data = []
                    for idx, row in df_skip.head(10).iterrows():
                        row_str = str(row.values)
                        if "202" in row_str or "主变" in row_str or "变压器" in row_str:
                            sample_data.append(f"第{idx+1}行: {dict(row)}")
                    if sample_data:
                        print(f"    包含关键词的行:")
                        for sample in sample_data[:2]:
                            print(f"      {sample}")
                break
        except Exception as e:
            continue
    
    # 方式3: 指定表头行
    for header_row in [0, 1, 2, 3]:
        try:
            df_header = pd.read_excel(file_path, header=header_row)
            if not df_header.empty and "202" in str(df_header.values) or "主变" in str(df_header.values):
                print(f"  方式3 (表头第{header_row}行): 形状 {df_header.shape}")
                print(f"    找到相关数据!")
                break
        except Exception as e:
            continue
    
    wb.close()

def find_target_data(file_path):
    """寻找目标数据"""
    print(f"\n🎯 寻找'202 2号主变'相关数据:")
    
    # 尝试多种读取方式
    for skip_rows in range(0, 8):
        try:
            df = pd.read_excel(file_path, skiprows=skip_rows)
            if df.empty:
                continue
                
            # 在所有列中搜索目标文本
            found_data = []
            for col in df.columns:
                col_data = df[col].astype(str)
                matches = col_data.str.contains("202.*2.*主变|2.*号.*主变", regex=True, na=False)
                if matches.any():
                    matched_rows = df[matches]
                    found_data.append({
                        'skip_rows': skip_rows,
                        'column': col,
                        'matches': len(matched_rows),
                        'sample_data': matched_rows.head(2).to_dict('records')
                    })
            
            if found_data:
                print(f"\n  ✅ 在跳过{skip_rows}行的数据中找到匹配:")
                for data in found_data:
                    print(f"    列 '{data['column']}': {data['matches']} 个匹配")
                    for sample in data['sample_data']:
                        print(f"      样本: {sample}")
                return skip_rows, found_data
                
        except Exception as e:
            continue
    
    print("  ❌ 未找到相关数据")
    return None, None

if __name__ == "__main__":
    file_path = "4月东宇母线不平衡率 .xlsx"
    if Path(file_path).exists():
        analyze_excel_structure(file_path)
        find_target_data(file_path)
    else:
        print(f"文件不存在: {file_path}")
