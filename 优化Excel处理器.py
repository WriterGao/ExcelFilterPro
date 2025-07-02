"""
优化的Excel处理器 - 专门处理复杂多表头电力报表
"""

import pandas as pd
import openpyxl
from typing import Dict, List, Tuple, Optional
import re

class PowerExcelProcessor:
    """电力行业Excel报表专用处理器"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.all_sheets_data = {}
        self.headers_info = {}
        
    def analyze_and_load(self):
        """分析并加载所有工作表"""
        print(f"🔍 分析文件: {self.file_path}")
        
        # 加载工作簿
        self.workbook = openpyxl.load_workbook(self.file_path)
        print(f"📋 发现 {len(self.workbook.sheetnames)} 个工作表")
        
        # 分析每个工作表
        for sheet_name in self.workbook.sheetnames:
            print(f"\n📊 处理工作表: {sheet_name}")
            try:
                # 智能检测表头和数据
                df = self._smart_read_sheet(sheet_name)
                if df is not None and not df.empty:
                    self.all_sheets_data[sheet_name] = df
                    print(f"✅ 成功加载 {len(df)} 行数据")
                else:
                    print("⚠️  工作表为空或无有效数据")
            except Exception as e:
                print(f"❌ 处理失败: {e}")
    
    def _smart_read_sheet(self, sheet_name: str) -> Optional[pd.DataFrame]:
        """智能读取工作表"""
        # 方法1：跳过前2行，第3行作为表头
        try:
            df = pd.read_excel(self.file_path, sheet_name=sheet_name, skiprows=2)
            
            # 清理列名
            df.columns = self._clean_column_names(df.columns)
            
            # 移除空行
            df = df.dropna(how='all')
            
            # 重置索引
            df = df.reset_index(drop=True)
            
            return df
            
        except Exception as e:
            print(f"智能读取失败: {e}")
            return None
    
    def _clean_column_names(self, columns):
        """清理列名"""
        cleaned_columns = []
        for i, col in enumerate(columns):
            if pd.isna(col) or str(col).startswith('Unnamed'):
                # 为未命名列生成有意义的名称
                if i == 0:
                    cleaned_columns.append('设备编号')
                elif i == 1:
                    cleaned_columns.append('上月电度表指示数')
                elif i == 2:
                    cleaned_columns.append('本月电度表指示数')
                elif i == 3:
                    cleaned_columns.append('倍率')
                elif i == 4:
                    cleaned_columns.append('输出电量')
                else:
                    cleaned_columns.append(f'列{i+1}')
            else:
                cleaned_columns.append(str(col).strip())
        
        return cleaned_columns
    
    def search_transformer_data(self, transformer_name: str, target_column: str):
        """搜索指定主变的数据"""
        print(f"\n🎯 搜索目标: {transformer_name} 的 {target_column}")
        print("="*60)
        
        results = {}
        
        for sheet_name, df in self.all_sheets_data.items():
            found_data = self._search_in_dataframe(df, transformer_name, target_column)
            if found_data:
                results[sheet_name] = found_data
                
        return results
    
    def _search_in_dataframe(self, df: pd.DataFrame, transformer_name: str, target_column: str):
        """在单个DataFrame中搜索数据"""
        found_rows = []
        
        # 在所有列中搜索变压器名称
        for col in df.columns:
            if df[col].dtype == 'object':  # 文本列
                mask = df[col].astype(str).str.contains(transformer_name, na=False, case=False)
                if mask.any():
                    matched_rows = df[mask]
                    for _, row in matched_rows.iterrows():
                        # 尝试找到目标列的值
                        target_value = self._extract_target_value(row, target_column)
                        if target_value is not None:
                            found_rows.append({
                                '设备名称': transformer_name,
                                '匹配列': col,
                                '目标值': target_value,
                                '完整行数据': row.to_dict()
                            })
        
        return found_rows
    
    def _extract_target_value(self, row, target_column: str):
        """提取目标列的值"""
        # 尝试多种匹配方式
        target_patterns = [
            target_column,
            '上月电度表指示数',
            '电度表',
            '指示数'
        ]
        
        for pattern in target_patterns:
            for col in row.index:
                if pattern in str(col):
                    value = row[col]
                    if pd.notna(value) and value != '':
                        return value
        
        # 如果没有找到匹配的列，返回第二列的值（通常是数值列）
        if len(row) > 1:
            return row.iloc[1]
        
        return None
    
    def get_summary(self):
        """获取处理摘要"""
        print(f"\n📈 处理摘要:")
        print(f"总工作表数: {len(self.workbook.sheetnames)}")
        print(f"成功处理: {len(self.all_sheets_data)}")
        print(f"失败: {len(self.workbook.sheetnames) - len(self.all_sheets_data)}")
        
        for sheet_name, df in self.all_sheets_data.items():
            print(f"  {sheet_name}: {len(df)} 行 x {len(df.columns)} 列")

def demo_power_excel_processing():
    """演示电力Excel处理"""
    file_path = "4月东宇母线不平衡率 .xlsx"
    
    # 创建处理器
    processor = PowerExcelProcessor(file_path)
    
    # 分析和加载
    processor.analyze_and_load()
    
    # 获取摘要
    processor.get_summary()
    
    # 搜索目标数据
    results = processor.search_transformer_data("202 2号主变", "上月电度表指示数")
    
    # 显示结果
    print(f"\n🎯 搜索结果:")
    if results:
        for sheet_name, sheet_results in results.items():
            print(f"\n📅 工作表: {sheet_name}")
            for i, result in enumerate(sheet_results):
                print(f"  结果 {i+1}:")
                print(f"    设备名称: {result['设备名称']}")
                print(f"    目标值: {result['目标值']}")
                print(f"    匹配列: {result['匹配列']}")
    else:
        print("❌ 未找到匹配数据")

if __name__ == "__main__":
    demo_power_excel_processing()
