"""
增强版Excel处理器 - 支持复杂多表头、多工作表的电力报表
"""

import pandas as pd
import openpyxl
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple
import numpy as np
from datetime import datetime
import string

from ..utils.logger import get_logger
from ..utils.helpers import validate_excel_file, format_file_size
from ..utils.exceptions import ExcelFileError, FileProcessingError
from ..utils.constants import SUPPORTED_EXCEL_FORMATS, MAX_DATA_ROWS


class EnhancedExcelProcessor:
    """增强版Excel文件处理核心类 - 支持复杂电力报表"""
    
    def __init__(self):
        self.logger = get_logger(__name__)
        self.data_frames: Dict[str, pd.DataFrame] = {}
        self.file_info: Dict[str, Dict[str, Any]] = {}
        self.headers: Dict[str, List[str]] = {}
        self.data_types: Dict[str, Dict[str, str]] = {}
        self.workbook_info: Dict[str, Any] = {}
        
    def load_excel_files(self, file_paths: List[str]) -> Dict[str, pd.DataFrame]:
        """加载多个Excel文件，支持复杂格式"""
        self.logger.info(f"开始加载{len(file_paths)}个Excel文件（增强模式）")
        loaded_files = {}
        
        for file_path in file_paths:
            try:
                validate_excel_file(file_path)
                
                file_key = Path(file_path).stem
                
                # 检测文件类型
                if self._is_complex_power_report(file_path):
                    self.logger.info(f"检测到复杂电力报表: {file_path}")
                    dfs = self._load_complex_power_report(file_path)
                    loaded_files.update(dfs)
                else:
                    # 使用标准方法加载
                    df = self._load_standard_excel(file_path)
                    if df is not None and not df.empty:
                        loaded_files[file_key] = df
                
                self.logger.info(f"成功加载文件: {file_path}")
                
            except Exception as e:
                self.logger.error(f"加载文件失败 {file_path}: {e}")
                raise ExcelFileError(f"无法加载Excel文件 {file_path}: {str(e)}")
        
        self.data_frames.update(loaded_files)
        self._extract_all_metadata()
        
        self.logger.info(f"成功加载{len(loaded_files)}个数据表")
        return loaded_files
    
    def _is_complex_power_report(self, file_path: str) -> bool:
        """检测是否为复杂的电力报表"""
        try:
            # 读取文件基本信息
            wb = openpyxl.load_workbook(file_path)
            
            # 检查工作表数量（多工作表通常是时间序列数据）
            if len(wb.sheetnames) > 5:
                return True
            
            # 检查第一个工作表的内容
            first_sheet = wb[wb.sheetnames[0]]
            
            # 检查前几行是否包含电力相关关键词
            power_keywords = ['母线', '主变', '变电站', '电度', '不平衡', 'kV']
            for row in range(1, min(6, first_sheet.max_row + 1)):
                for col in range(1, min(6, first_sheet.max_column + 1)):
                    cell_value = str(first_sheet.cell(row=row, column=col).value or '')
                    if any(keyword in cell_value for keyword in power_keywords):
                        return True
            
            return False
        except:
            return False
    
    def _load_complex_power_report(self, file_path: str) -> Dict[str, pd.DataFrame]:
        """加载复杂的电力报表"""
        results = {}
        file_key = Path(file_path).stem
        
        try:
            wb = openpyxl.load_workbook(file_path)
            
            # 记录工作簿信息
            self.workbook_info[file_key] = {
                'path': file_path,
                'sheet_count': len(wb.sheetnames),
                'sheet_names': wb.sheetnames
            }
            
            # 处理每个工作表
            for sheet_name in wb.sheetnames:
                try:
                    df = self._smart_read_sheet(file_path, sheet_name)
                    if df is not None and not df.empty:
                        # 组合键：文件名_工作表名
                        combined_key = f"{file_key}_{sheet_name}"
                        results[combined_key] = df
                        self.logger.info(f"  加载工作表 {sheet_name}: {len(df)} 行")
                except Exception as e:
                    self.logger.warning(f"  工作表 {sheet_name} 加载失败: {e}")
                    
        except Exception as e:
            self.logger.error(f"复杂报表加载失败: {e}")
            
        return results
    
    def _smart_read_sheet(self, file_path: str, sheet_name: str) -> Optional[pd.DataFrame]:
        """智能读取工作表，自动处理表头"""
        try:
            # 尝试不同的skiprows值
            for skip_rows in [2, 1, 3, 0]:
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=skip_rows)
                    
                    if df.empty:
                        continue
                    
                    # 清理列名
                    df.columns = self._clean_column_names(df.columns)
                    
                    # 移除空行
                    df = df.dropna(how='all').reset_index(drop=True)
                    
                    # 检查是否有有效数据
                    if len(df) > 0 and any(not pd.isna(val) for val in df.iloc[0]):
                        return df
                        
                except Exception:
                    continue
            
            return None
            
        except Exception as e:
            self.logger.error(f"智能读取工作表失败 {sheet_name}: {e}")
            return None
    
    def _load_standard_excel(self, file_path: str) -> Optional[pd.DataFrame]:
        """标准Excel文件加载方法"""
        try:
            # 使用header=None保持原始行结构，不把第一行当作列名
            df = pd.read_excel(file_path, engine='openpyxl', header=None)
            
            # 生成默认列名 A, B, C, D...
            if not df.empty:
                max_cols = len(df.columns)
                if max_cols <= 26:
                    df.columns = list(string.ascii_uppercase)[:max_cols]
                else:
                    # 处理超过26列的情况：A, B, ..., Z, AA, AB, ...
                    columns = []
                    for i in range(max_cols):
                        if i < 26:
                            columns.append(string.ascii_uppercase[i])
                        else:
                            first = string.ascii_uppercase[(i-26)//26]
                            second = string.ascii_uppercase[i%26] 
                            columns.append(first + second)
                    df.columns = columns
            
            df = self._clean_dataframe(df)
            return df
        except Exception as e:
            self.logger.error(f"标准Excel加载失败: {e}")
            return None
    
    def _clean_column_names(self, columns):
        """智能清理列名"""
        cleaned_columns = []
        
        for i, col in enumerate(columns):
            col_str = str(col) if col is not None else "None"
            col_str = col_str.strip()
            
            if pd.isna(col) or col_str.startswith('Unnamed') or col_str == 'nan':
                # 根据位置推断列名
                if i == 0:
                    cleaned_columns.append('设备编号')
                elif i == 1:
                    cleaned_columns.append('上月电度表指示数')
                elif i == 2:
                    cleaned_columns.append('本月电度表指示数')
                elif i == 3:
                    cleaned_columns.append('倍率')
                elif i == 4:
                    cleaned_columns.append('输出电量')
                else:
                    cleaned_columns.append(f'列{i+1}')
            else:
                # 保留原列名，但做基本清理
                clean_name = col_str.replace('\n', '').replace('\r', '').strip()
                if not clean_name:
                    clean_name = f'列{i+1}'
                cleaned_columns.append(clean_name)
        
        return cleaned_columns
    
    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """清理DataFrame数据"""
        # 只移除完全为空的行，保留所有列（包括部分为空的列）
        df = df.dropna(how='all')
        
        # 重置索引
        df = df.reset_index(drop=True)
        
        # 处理列名
        df.columns = self._clean_column_names(df.columns)
        
        return df
    
    def search_across_all_sheets(self, search_term: str, target_column: str = None) -> Dict[str, List[Dict]]:
        """跨所有工作表搜索数据"""
        results = {}
        
        for sheet_key, df in self.data_frames.items():
            sheet_results = self._search_in_dataframe(df, search_term, target_column)
            if sheet_results:
                results[sheet_key] = sheet_results
        
        return results
    
    def _search_in_dataframe(self, df: pd.DataFrame, search_term: str, target_column: str = None):
        """在DataFrame中搜索"""
        found_rows = []
        
        # 在所有文本列中搜索
        for col in df.columns:
            if df[col].dtype == 'object':
                mask = df[col].astype(str).str.contains(search_term, na=False, case=False)
                if mask.any():
                    matched_rows = df[mask]
                    for idx, row in matched_rows.iterrows():
                        target_value = self._extract_target_value(row, target_column)
                        found_rows.append({
                            '行索引': idx,
                            '匹配列': col,
                            '匹配值': row[col],
                            '目标值': target_value,
                            '完整行': row.to_dict()
                        })
        
        return found_rows
    
    def _extract_target_value(self, row, target_column: str = None):
        """提取目标列的值"""
        if target_column:
            # 精确匹配
            if target_column in row.index:
                return row[target_column]
            
            # 模糊匹配
            for col in row.index:
                if target_column.lower() in str(col).lower():
                    return row[col]
        
        # 默认返回第二列（通常是数值列）
        if len(row) > 1:
            return row.iloc[1]
        
        return None
    
    def get_all_unique_headers(self) -> List[str]:
        """获取所有文件的唯一表头"""
        all_headers = set()
        for headers in self.headers.values():
            all_headers.update(headers)
        
        return sorted(list(all_headers))
    
    def extract_headers(self, df: pd.DataFrame) -> List[str]:
        """提取DataFrame的表头"""
        return df.columns.tolist()
    
    def infer_data_types(self, df: pd.DataFrame) -> Dict[str, str]:
        """推断DataFrame各列的数据类型"""
        type_mapping = {}
        
        for column in df.columns:
            col_data = df[column].dropna()
            
            if len(col_data) == 0:
                type_mapping[column] = 'string'
                continue
            
            if pd.api.types.is_numeric_dtype(col_data):
                type_mapping[column] = 'number'
            elif pd.api.types.is_datetime64_any_dtype(col_data):
                type_mapping[column] = 'datetime'
            elif pd.api.types.is_bool_dtype(col_data):
                type_mapping[column] = 'boolean'
            else:
                type_mapping[column] = 'string'
        
        return type_mapping
    
    def _extract_all_metadata(self):
        """提取所有已加载文件的元数据"""
        for file_key, df in self.data_frames.items():
            self.headers[file_key] = self.extract_headers(df)
            self.data_types[file_key] = self.infer_data_types(df)
    
    def get_combined_dataframe(self) -> pd.DataFrame:
        """将所有数据源合并为一个DataFrame"""
        if not self.data_frames:
            return pd.DataFrame()
        
        all_columns = self.get_all_unique_headers()
        standardized_dfs = []
        
        for file_key, df in self.data_frames.items():
            # 为每个DataFrame添加缺失的列
            for col in all_columns:
                if col not in df.columns:
                    df[col] = None
            
            # 重新排序列
            df = df[all_columns]
            
            # 添加数据源标识
            df['_source_file'] = file_key
            
            standardized_dfs.append(df)
        
        combined_df = pd.concat(standardized_dfs, ignore_index=True)
        return combined_df
    
    def get_file_summary(self) -> Dict[str, Any]:
        """获取已加载文件的汇总信息"""
        summary = {
            'total_files': len(self.data_frames),
            'total_rows': sum(len(df) for df in self.data_frames.values()),
            'total_columns': len(self.get_all_unique_headers()),
            'files_info': [],
            'workbook_info': self.workbook_info
        }
        
        for file_key, df in self.data_frames.items():
            summary['files_info'].append({
                'name': file_key,
                'rows': len(df),
                'columns': len(df.columns),
                'headers': self.headers.get(file_key, [])
            })
        
        return summary
    
    def write_to_template(self, template_path: str, filtered_data: Dict[str, pd.DataFrame], 
                         output_path: str) -> bool:
        """
        将筛选结果写入模板文件
        
        Args:
            template_path: 模板文件路径
            filtered_data: 筛选结果数据，格式为{列名: DataFrame}
            output_path: 输出文件路径
            
        Returns:
            是否写入成功
        """
        try:
            # 加载模板文件
            template_df = self._load_standard_excel(template_path)
            if template_df is None:
                raise ExcelFileError("无法加载模板文件")
            
            # 复制模板结构
            output_df = template_df.copy()
            
            # 清空数据行，保留表头
            if len(output_df) > 0:
                output_df = output_df.iloc[0:0]  # 保留列结构，清空行
            
            # 计算需要的总行数
            max_rows_needed = 0
            for column_name, data in filtered_data.items():
                if column_name in output_df.columns:
                    max_rows_needed = max(max_rows_needed, len(data))
            
            # 扩展输出DataFrame的行数
            if max_rows_needed > 0:
                # 创建空行
                empty_rows = pd.DataFrame(index=range(max_rows_needed), 
                                        columns=output_df.columns)
                output_df = pd.concat([output_df, empty_rows], ignore_index=True)
            
            # 填充筛选结果到对应列
            for column_name, data in filtered_data.items():
                if column_name in output_df.columns and not data.empty:
                    # 取第一列的数据（假设筛选结果是单列数据或取第一个非空列）
                    if len(data.columns) > 0:
                        source_column = data.columns[0]
                        values = data[source_column].dropna().tolist()
                        
                        # 填充到目标列
                        for i, value in enumerate(values):
                            if i < len(output_df):
                                output_df.loc[i, column_name] = value
            
            # 保存到输出文件
            output_df.to_excel(output_path, index=False, engine='openpyxl')
            
            self.logger.info(f"成功将结果写入文件: {output_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"写入模板文件失败: {e}")
            raise FileProcessingError(f"写入模板文件失败: {str(e)}")

    def clear_data(self):
        """清空所有加载的数据"""
        self.data_frames.clear()
        self.file_info.clear()
        self.headers.clear()
        self.data_types.clear()
        self.workbook_info.clear()
        self.logger.info("已清空所有数据")

    def get_original_file_names(self) -> List[str]:
        """获取原始文件名列表（不包含工作表后缀）"""
        original_names = set()
        
        # 从workbook_info中获取原始文件名
        for file_key, info in self.workbook_info.items():
            if 'path' in info:
                original_name = Path(info['path']).name
                original_names.add(original_name)
        
        # 如果workbook_info为空，从data_frames键中推断
        if not original_names:
            for data_key in self.data_frames.keys():
                # 移除工作表后缀（格式：文件名_工作表名）
                if '_' in data_key:
                    original_name = data_key.split('_')[0] + '.xlsx'
                else:
                    original_name = data_key + '.xlsx'
                original_names.add(original_name)
        
        return sorted(list(original_names))
    
    def get_sheets_for_file(self, original_file_name: str) -> List[str]:
        """获取指定原始文件的所有工作表数据源键"""
        file_stem = Path(original_file_name).stem
        matching_keys = []
        
        for data_key in self.data_frames.keys():
            if data_key.startswith(file_stem + '_'):
                matching_keys.append(data_key)
        
        return matching_keys
    
    def get_dataframe_by_original_name(self, original_file_name: str, sheet_hint: str = None) -> Optional[pd.DataFrame]:
        """根据原始文件名获取DataFrame"""
        sheets = self.get_sheets_for_file(original_file_name)
        
        if not sheets:
            return None
        
        # 如果指定了工作表提示，尝试匹配
        if sheet_hint:
            for sheet_key in sheets:
                if sheet_hint.lower() in sheet_key.lower():
                    return self.data_frames.get(sheet_key)
        
        # 否则返回第一个工作表的数据
        return self.data_frames.get(sheets[0])
